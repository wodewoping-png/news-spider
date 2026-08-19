from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.load_sources import Source
from src.rionews import split_rionews_workbook_by_date
from src.scrapers import get_scraper_class
from src.scrapers.rionews import (
    RIONewsBatteryScraper,
    RIONewsChinaEnergyScraper,
    RIONewsInternationalEnergyScraper,
    RIONewsXEVCarScraper,
)


HEADERS = [
    "新闻标题",
    "发布时间",
    "发布媒体",
    "链接",
    "发布企业",
    "新闻全文",
    "技术标签",
    "新闻标签",
    "关键词Sheet",
]


def source(name: str = "中国能源网") -> Source:
    return Source(
        name=name,
        media_type="综合能源媒体",
        domain="综合新能源",
        sub_domain="能源经济",
        frequency="工作日",
        description="",
        note="",
        url="https://www.china5e.com/news/",
    )


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "新闻命中标签导出"
    sheet.append(HEADERS)
    sheet.append(["新能源项目投运", "2026-08-04 09:30:00", "中国能源网_资讯_新能源", "https://www.china5e.com/news/news-1-1.html", "", "完整正文一" * 100, "", "", ""])
    sheet.append(["节能政策发布", "2026-08-04 10:30:00", "中国能源网_资讯_节能低碳", "https://www.china5e.com/news/news-2-1.html", "", "完整正文二" * 100, "", "", ""])
    sheet.append(["电池企业动态", "2026-08-04 10:40:00", "电池网_新闻_企业", "https://www.itdcw.com/news/qiye/1.html", "", "电池网正文" * 100, "", "", ""])
    sheet.append(["电动车动态", "2026-08-04 10:50:00", "我爱电车网_首页", "https://www.xevcar.com/gongsi/1.html", "", "我爱电车网正文" * 100, "", "", ""])
    sheet.append(["能源宏观动态", "2026-08-04 11:00:00", "国际能源网_宏观", "https://www.in-en.com/article/1.html", "", "国际能源网正文" * 100, "", "", ""])
    sheet.append(["其他媒体文章", "2026-08-04 11:30:00", "储能网", "https://example.com/other", "", "不应导入", "", "", ""])
    sheet.append(["其他日期文章", "2026-08-03 09:30:00", "中国能源网_资讯_新能源", "https://www.china5e.com/news/news-3-1.html", "", "不应导入", "", "", ""])
    workbook.save(path)
    workbook.close()


class RIONewsTests(unittest.TestCase):
    def test_missing_daily_workbook_falls_back_without_input_failure(self):
        class EmptyClient:
            def get(self, _url, **_kwargs):
                return None

        with tempfile.TemporaryDirectory() as temp:
            previous = os.environ.get("RIONEWS_DAILY_DIR")
            os.environ["RIONEWS_DAILY_DIR"] = temp
            try:
                scraper = RIONewsChinaEnergyScraper(EmptyClient(), source())
                articles = scraper.scrape(target_date=date(2026, 8, 12))
            finally:
                if previous is None:
                    os.environ.pop("RIONEWS_DAILY_DIR", None)
                else:
                    os.environ["RIONEWS_DAILY_DIR"] = previous

        self.assertEqual(articles, [])
        self.assertEqual(scraper.last_candidate_count, 0)
        self.assertEqual(scraper.last_fetched_count, 0)

    def test_registry_uses_rionews_for_china_energy(self):
        self.assertIs(get_scraper_class("中国能源网"), RIONewsChinaEnergyScraper)

    def test_registry_uses_rionews_for_three_replacement_sources(self):
        expected = {
            "电池网": RIONewsBatteryScraper,
            "我爱电车网": RIONewsXEVCarScraper,
            "国际能源网": RIONewsInternationalEnergyScraper,
        }
        for name, scraper_class in expected.items():
            with self.subTest(name=name):
                self.assertIs(get_scraper_class(name), scraper_class)

    def test_imports_only_china_energy_media_for_target_date(self):
        with tempfile.TemporaryDirectory() as temp:
            daily_dir = Path(temp)
            write_workbook(daily_dir / "news_export_2026-08-04.xlsx")
            previous = os.environ.get("RIONEWS_DAILY_DIR")
            os.environ["RIONEWS_DAILY_DIR"] = str(daily_dir)
            try:
                scraper = RIONewsChinaEnergyScraper(object(), source())
                articles = scraper.scrape(limit=1, target_date=date(2026, 8, 4), candidate_limit=100)
            finally:
                if previous is None:
                    os.environ.pop("RIONEWS_DAILY_DIR", None)
                else:
                    os.environ["RIONEWS_DAILY_DIR"] = previous

        self.assertEqual(len(articles), 2)
        self.assertEqual(scraper.last_candidate_count, 2)
        self.assertTrue(all(item["source_name"] == "中国能源网" for item in articles))
        self.assertTrue(all(item["domain"] == "综合新能源" for item in articles))
        self.assertTrue(all(item["published_at"].startswith("2026-08-04") for item in articles))

    def test_each_replacement_source_uses_only_its_exact_media_prefix(self):
        cases = (
            ("电池网", RIONewsBatteryScraper, "itdcw.com"),
            ("我爱电车网", RIONewsXEVCarScraper, "xevcar.com"),
            ("国际能源网", RIONewsInternationalEnergyScraper, "in-en.com"),
        )
        with tempfile.TemporaryDirectory() as temp:
            daily_dir = Path(temp)
            write_workbook(daily_dir / "news_export_2026-08-04.xlsx")
            previous = os.environ.get("RIONEWS_DAILY_DIR")
            os.environ["RIONEWS_DAILY_DIR"] = str(daily_dir)
            try:
                for name, scraper_class, domain in cases:
                    with self.subTest(name=name):
                        scraper = scraper_class(object(), source(name))
                        articles = scraper.scrape(target_date=date(2026, 8, 4), candidate_limit=100)
                        self.assertEqual(len(articles), 1)
                        self.assertEqual(articles[0]["source_name"], name)
                        self.assertIn(domain, articles[0]["url"])
            finally:
                if previous is None:
                    os.environ.pop("RIONEWS_DAILY_DIR", None)
                else:
                    os.environ["RIONEWS_DAILY_DIR"] = previous

    def test_split_export_creates_daily_workbooks(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            aggregate = root / "news_export.xlsx"
            daily_dir = root / "daily"
            write_workbook(aggregate)
            outputs = split_rionews_workbook_by_date(aggregate, daily_dir, required_date=date(2026, 8, 5))
            self.assertEqual(
                {path.name for path in outputs},
                {"news_export_2026-08-03.xlsx", "news_export_2026-08-04.xlsx", "news_export_2026-08-05.xlsx"},
            )
            empty = load_workbook(daily_dir / "news_export_2026-08-05.xlsx", read_only=True)
            try:
                self.assertEqual(empty.active.max_row, 1)
            finally:
                empty.close()


if __name__ == "__main__":
    unittest.main()

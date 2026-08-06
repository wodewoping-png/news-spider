from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from .date_utils import DEFAULT_TIMEZONE
from .load_sources import Source


CHINA_ENERGY_MEDIA_PREFIX = "中国能源网_"
HEADER_ALIASES = {
    "title": {"新闻标题", "title"},
    "published_at": {"发布时间", "publish_time", "publication_time", "date"},
    "media": {"发布媒体", "source", "source_name", "media"},
    "url": {"链接", "url", "link"},
    "company": {"发布企业", "company"},
    "content": {"新闻全文", "content", "body"},
    "technical_tags": {"技术标签", "technical_tags"},
    "news_tags": {"新闻标签", "news_tags"},
}


def _header_variants(value: Any) -> set[str]:
    text = str(value or "").strip()
    variants = {text.casefold()}
    for encoding in ("latin1", "cp1252"):
        try:
            repaired = text.encode(encoding).decode("gbk").strip()
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
        variants.add(repaired.casefold())
    return variants


def _column_map(headers: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, header in enumerate(headers):
        variants = _header_variants(header)
        for field, aliases in HEADER_ALIASES.items():
            if field not in result and variants & {alias.casefold() for alias in aliases}:
                result[field] = index
    required = {"title", "published_at", "media", "url", "content"}
    missing = sorted(required - result.keys())
    if missing:
        raise ValueError(f"RIOnews workbook is missing columns: {', '.join(missing)}")
    return result


def parse_rionews_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("/", "-"))
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=ZoneInfo(DEFAULT_TIMEZONE))
    return parsed.astimezone(ZoneInfo(DEFAULT_TIMEZONE))


def load_rionews_articles(
    workbook_path: Path,
    source: Source,
    *,
    target_date: date | None = None,
    media_prefix: str = CHINA_ENERGY_MEDIA_PREFIX,
) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []
        columns = _column_map(headers)
        articles: list[dict] = []
        seen_urls: set[str] = set()
        for row in rows:
            media = str(row[columns["media"]] or "").strip()
            if not media.startswith(media_prefix):
                continue
            published = parse_rionews_datetime(row[columns["published_at"]])
            if not published or (target_date and published.date() != target_date):
                continue
            title = str(row[columns["title"]] or "").strip()
            url = str(row[columns["url"]] or "").strip()
            content = str(row[columns["content"]] or "").strip()
            if not title or not url or url in seen_urls:
                continue
            seen_urls.add(url)
            articles.append(
                {
                    "title": title,
                    "published_at": published.isoformat(),
                    "content": content,
                    "url": url,
                    "source_name": source.name,
                    "domain": source.domain,
                    "sub_domain": source.sub_domain,
                    "crawled_at": datetime.now(timezone.utc).isoformat(),
                }
            )
        return articles
    finally:
        workbook.close()


def split_rionews_workbook_by_date(
    workbook_path: Path,
    output_dir: Path,
    *,
    required_date: date | None = None,
) -> list[Path]:
    source = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sheet = source.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("RIOnews workbook is empty")
        columns = _column_map(headers)
        rows_by_date: dict[date, list[tuple[Any, ...]]] = {}
        for row in rows:
            published = parse_rionews_datetime(row[columns["published_at"]])
            if published:
                rows_by_date.setdefault(published.date(), []).append(tuple(row))
        if required_date is not None:
            rows_by_date.setdefault(required_date, [])
        if not rows_by_date:
            raise ValueError("RIOnews workbook has no dated rows")

        output_dir.mkdir(parents=True, exist_ok=True)
        outputs: list[Path] = []
        for publish_date, dated_rows in sorted(rows_by_date.items()):
            daily = Workbook()
            daily_sheet = daily.active
            daily_sheet.title = sheet.title
            daily_sheet.append(list(headers))
            for row in dated_rows:
                daily_sheet.append(list(row))
            daily_sheet.freeze_panes = "A2"
            daily_sheet.auto_filter.ref = daily_sheet.dimensions
            output = output_dir / f"news_export_{publish_date.isoformat()}.xlsx"
            temporary = output.with_suffix(".xlsx.part")
            try:
                daily.save(temporary)
                temporary.replace(output)
            finally:
                daily.close()
                if temporary.exists():
                    temporary.unlink()
            outputs.append(output)
        return outputs
    finally:
        source.close()

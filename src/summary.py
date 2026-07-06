from __future__ import annotations

import argparse
import calendar
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Sequence
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .date_utils import DEFAULT_TIMEZONE
from .storage import FIELDNAMES


SUMMARY_FIELDNAMES = tuple(FIELDNAMES) + ("category", "category_method")
DAILY_FILE_RE = re.compile(r"articles-(?P<date>\d{4}-\d{2}-\d{2})\.csv$")
WEEK_RE = re.compile(r"(?P<year>\d{4})-W(?P<week>\d{2})$")


def _local_today(timezone_name: str) -> date:
    return datetime.now(ZoneInfo(timezone_name)).date()


def previous_week_bounds(timezone_name: str = DEFAULT_TIMEZONE) -> tuple[date, date]:
    today = _local_today(timezone_name)
    this_monday = today - timedelta(days=today.weekday())
    start = this_monday - timedelta(days=7)
    return start, start + timedelta(days=6)


def week_bounds(week: str) -> tuple[date, date]:
    match = WEEK_RE.fullmatch(week)
    if not match:
        raise ValueError("week must use YYYY-Www format, for example 2026-W23")
    year = int(match.group("year"))
    week_number = int(match.group("week"))
    start = date.fromisocalendar(year, week_number, 1)
    return start, start + timedelta(days=6)


def previous_month(timezone_name: str = DEFAULT_TIMEZONE) -> str:
    today = _local_today(timezone_name)
    first_this_month = today.replace(day=1)
    last_previous_month = first_this_month - timedelta(days=1)
    return last_previous_month.strftime("%Y-%m")


def month_bounds(month: str) -> tuple[date, date]:
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise ValueError("month must use YYYY-MM format")
    year, month_number = map(int, month.split("-"))
    last_day = calendar.monthrange(year, month_number)[1]
    return date(year, month_number, 1), date(year, month_number, last_day)


def iter_dates(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def daily_csv_path(data_dir: Path, day: date) -> Path:
    return data_dir / f"articles-{day.isoformat()}.csv"


def _empty_articles() -> pd.DataFrame:
    return pd.DataFrame(columns=list(SUMMARY_FIELDNAMES) + ["summary_date"])


def _read_daily_csv(path: Path, file_date: date) -> pd.DataFrame:
    try:
        df = pd.read_csv(path, keep_default_na=False, encoding="utf-8-sig")
    except UnicodeDecodeError:
        df = pd.read_csv(path, keep_default_na=False, encoding="utf-8")

    for column in SUMMARY_FIELDNAMES:
        if column not in df.columns:
            df[column] = ""

    df = df.loc[:, list(SUMMARY_FIELDNAMES)].copy()
    df["summary_date"] = file_date
    return df


def load_articles_between(data_dir: Path, start: date, end: date) -> tuple[pd.DataFrame, list[Path]]:
    frames: list[pd.DataFrame] = []
    files: list[Path] = []

    for day in iter_dates(start, end):
        path = daily_csv_path(data_dir, day)
        if not path.exists():
            continue
        frames.append(_read_daily_csv(path, day))
        files.append(path)

    if not frames:
        return _empty_articles(), files

    articles = pd.concat(frames, ignore_index=True)
    articles = normalize_articles(articles, start, end)
    return articles, files


def normalize_articles(df: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    if df.empty:
        return _empty_articles()

    normalized = df.copy()
    for column in SUMMARY_FIELDNAMES:
        if column not in normalized.columns:
            normalized[column] = ""
        normalized[column] = normalized[column].fillna("").astype(str).str.strip()

    published = pd.to_datetime(normalized["published_at"], errors="coerce")
    normalized["summary_date"] = published.dt.date.where(
        published.notna(),
        normalized["summary_date"],
    )
    normalized = normalized[
        (normalized["summary_date"] >= start) & (normalized["summary_date"] <= end)
    ].copy()
    if normalized.empty:
        return _empty_articles()

    normalized["title_key"] = (
        normalized["title"].fillna("").astype(str).str.lower().str.replace(r"\s+", " ", regex=True).str.strip()
    )
    normalized["dedup_key"] = normalized["url"].where(
        normalized["url"].astype(str).str.len() > 0,
        normalized["source_name"].astype(str) + "::" + normalized["title_key"],
    )
    normalized = normalized.sort_values(
        ["summary_date", "source_name", "title"],
        kind="stable",
    )
    normalized = normalized.drop_duplicates(subset=["dedup_key"], keep="first")
    return normalized.loc[:, list(SUMMARY_FIELDNAMES) + ["summary_date"]]


def _format_date(value: object) -> str:
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def detail_for_export(articles: pd.DataFrame) -> pd.DataFrame:
    if articles.empty:
        return pd.DataFrame(columns=list(SUMMARY_FIELDNAMES))
    out = articles.loc[:, list(SUMMARY_FIELDNAMES)].copy()
    if "published_at" in out.columns:
        out["published_at"] = articles["summary_date"].map(_format_date)
    return out


def summarize_by_source(articles: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "source_name",
        "domain",
        "sub_domain",
        "category",
        "count",
        "first_date",
        "last_date",
    ]
    if articles.empty:
        return pd.DataFrame(columns=columns)

    grouped = (
        articles.groupby(["source_name", "domain", "sub_domain", "category"], dropna=False)
        .agg(
            count=("title", "size"),
            first_date=("summary_date", "min"),
            last_date=("summary_date", "max"),
        )
        .reset_index()
    )
    grouped["first_date"] = grouped["first_date"].map(_format_date)
    grouped["last_date"] = grouped["last_date"].map(_format_date)
    return grouped.sort_values(["count", "source_name"], ascending=[False, True], kind="stable")


def summarize_by_domain(articles: pd.DataFrame) -> pd.DataFrame:
    columns = ["domain", "sub_domain", "count"]
    if articles.empty:
        return pd.DataFrame(columns=columns)
    grouped = (
        articles.groupby(["domain", "sub_domain"], dropna=False)
        .size()
        .reset_index(name="count")
    )
    return grouped.sort_values(["count", "domain", "sub_domain"], ascending=[False, True, True], kind="stable")


def summarize_by_day(articles: pd.DataFrame) -> pd.DataFrame:
    columns = ["date", "count"]
    if articles.empty:
        return pd.DataFrame(columns=columns)
    grouped = articles.groupby("summary_date", dropna=False).size().reset_index(name="count")
    grouped["date"] = grouped["summary_date"].map(_format_date)
    return grouped.loc[:, columns].sort_values("date", kind="stable")


def summarize_by_category(articles: pd.DataFrame) -> pd.DataFrame:
    columns = ["category", "category_method", "count"]
    if articles.empty:
        return pd.DataFrame(columns=columns)
    grouped = (
        articles.groupby(["category", "category_method"], dropna=False)
        .size()
        .reset_index(name="count")
    )
    return grouped.sort_values(["count", "category"], ascending=[False, True], kind="stable")


def write_weekly_csv(
    data_dir: Path,
    output_dir: Path,
    start: date,
    end: date,
) -> Path:
    articles, files = load_articles_between(data_dir, start, end)
    output_dir.mkdir(parents=True, exist_ok=True)
    iso = end.isocalendar()
    output_path = output_dir / f"articles_week_{iso.year}-W{iso.week:02d}.csv"
    detail_for_export(articles).to_csv(output_path, index=False, encoding="utf-8-sig")
    print(f"[weekly] Window: {start} to {end}")
    print(f"[weekly] Read {len(files)} daily file(s), wrote {len(articles)} record(s): {output_path}")
    return output_path


def _autosize_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    header_to_col = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    url_col = header_to_col.get("url")
    if url_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=url_col)
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def write_monthly_workbook(
    data_dir: Path,
    output_dir: Path,
    month: str,
) -> Path:
    start, end = month_bounds(month)
    articles, files = load_articles_between(data_dir, start, end)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"articles_month_{month}.xlsx"

    sheets = {
        "Articles": detail_for_export(articles),
        "Source Summary": summarize_by_source(articles),
        "Domain Summary": summarize_by_domain(articles),
        "Daily Summary": summarize_by_day(articles),
        "Category Placeholder": summarize_by_category(articles),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

        for worksheet in writer.book.worksheets:
            _autosize_sheet(worksheet)

    print(f"[monthly] Window: {start} to {end}")
    print(f"[monthly] Read {len(files)} daily file(s), wrote {len(articles)} record(s): {output_path}")
    return output_path


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Aggregate daily news CSV files.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    weekly = subparsers.add_parser("weekly", help="Build a weekly CSV summary.")
    weekly.add_argument("--data-dir", type=Path, default=Path("data"))
    weekly.add_argument("--output-dir", type=Path, default=Path("data") / "weekly")
    weekly.add_argument("--week", help="ISO week in YYYY-Www format. Defaults to previous full week.")
    weekly.add_argument("--start-date", help="Custom window start date, YYYY-MM-DD.")
    weekly.add_argument("--end-date", help="Custom window end date, YYYY-MM-DD.")
    weekly.add_argument("--timezone", default=DEFAULT_TIMEZONE)

    monthly = subparsers.add_parser("monthly", help="Build a monthly Excel summary.")
    monthly.add_argument("--data-dir", type=Path, default=Path("data"))
    monthly.add_argument("--output-dir", type=Path, default=Path("data") / "monthly")
    monthly.add_argument("--month", help="Target month in YYYY-MM. Defaults to previous month.")
    monthly.add_argument("--timezone", default=DEFAULT_TIMEZONE)

    return parser.parse_args(argv)


def weekly_bounds_from_args(args: argparse.Namespace) -> tuple[date, date]:
    if args.start_date or args.end_date:
        if not args.start_date or not args.end_date:
            raise ValueError("--start-date and --end-date must be used together")
        start = _parse_date(args.start_date)
        end = _parse_date(args.end_date)
    elif args.week:
        start, end = week_bounds(args.week)
    else:
        start, end = previous_week_bounds(args.timezone)

    if start > end:
        raise ValueError("start date must be before or equal to end date")
    return start, end


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)

    if args.command == "weekly":
        start, end = weekly_bounds_from_args(args)
        write_weekly_csv(args.data_dir, args.output_dir, start, end)
        return 0

    if args.command == "monthly":
        month = args.month or previous_month(args.timezone)
        write_monthly_workbook(args.data_dir, args.output_dir, month)
        return 0

    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
